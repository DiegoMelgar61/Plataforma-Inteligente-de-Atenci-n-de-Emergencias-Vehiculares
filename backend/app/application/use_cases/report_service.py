"""
Generación de reportes personalizados (xlsx / pdf) a partir de búsquedas.
Respeta el filtrado por tenant. Determinístico, sin estado.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

from sqlalchemy.orm import Session

from app.models.models import ASIGNACIONES, INCIDENTES, TALLERES


def _val(v) -> str:
    return v.value if hasattr(v, "value") else ("" if v is None else str(v))


def construir_reporte_incidentes(
    db: Session,
    *,
    tenant_efectivo=None,
    estado: str | None = None,
    clasificacion: str | None = None,
    desde: datetime | None = None,
    hasta: datetime | None = None,
) -> tuple[list[str], list[list[str]]]:
    """
    Devuelve (columnas, filas) de incidentes según los filtros.
    `tenant_efectivo=None` significa sin filtro de tenant (solo ADMIN sin filtrar).
    """
    q = (
        db.query(INCIDENTES, TALLERES.NOMBRE_NEGOCIO, ASIGNACIONES.MONTO_COTIZADO)
        .outerjoin(
            ASIGNACIONES,
            (ASIGNACIONES.ID_INCIDENTE == INCIDENTES.ID_INCIDENTE)
            & (ASIGNACIONES.FECHA_RECHAZO.is_(None)),
        )
        .outerjoin(TALLERES, TALLERES.ID_TALLER == ASIGNACIONES.ID_TALLER)
    )
    if tenant_efectivo:
        q = q.filter(INCIDENTES.ID_TENANT == tenant_efectivo)
    if estado:
        q = q.filter(INCIDENTES.ESTADO == estado)
    if clasificacion:
        q = q.filter(INCIDENTES.CLASIFICACION == clasificacion)
    if desde:
        q = q.filter(INCIDENTES.FECHA_CREACION >= desde)
    if hasta:
        q = q.filter(INCIDENTES.FECHA_CREACION <= hasta)

    filas_orm = q.order_by(INCIDENTES.FECHA_CREACION.desc()).all()

    columnas = [
        "ID", "Fecha", "Estado", "Prioridad", "Clasificación",
        "Taller asignado", "Monto (Bs.)", "Resumen IA",
    ]
    filas: list[list[str]] = []
    for inc, taller_nombre, monto in filas_orm:
        fecha = inc.FECHA_CREACION.strftime("%Y-%m-%d %H:%M") if inc.FECHA_CREACION else ""
        resumen = (inc.RESUMEN_IA or "").replace("\n", " ").strip()
        if len(resumen) > 180:
            resumen = resumen[:177] + "..."
        filas.append([
            str(inc.ID_INCIDENTE)[:8].upper(),
            fecha,
            _val(inc.ESTADO),
            _val(inc.PRIORIDAD),
            _val(inc.CLASIFICACION),
            taller_nombre or "—",
            f"{float(monto):.2f}" if monto is not None else "—",
            resumen or "—",
        ])
    return columnas, filas


def generar_xlsx(titulo: str, columnas: list[str], filas: list[list[str]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append([titulo])
    ws.append([f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}  ·  {len(filas)} registros"])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(columnas)
    header_fill = PatternFill("solid", fgColor="4F46E5")
    for col_idx in range(1, len(columnas) + 1):
        c = ws.cell(row=header_row, column=col_idx)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for fila in filas:
        ws.append(fila)

    anchos = [12, 18, 14, 12, 16, 26, 14, 60]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = ancho

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_pdf(titulo: str, columnas: list[str], filas: list[list[str]]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24,
    )
    estilos = getSampleStyleSheet()
    celda = estilos["BodyText"]
    celda.fontSize = 7
    celda.leading = 9

    elementos = [
        Paragraph(titulo, estilos["Title"]),
        Paragraph(
            f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')} · {len(filas)} registros",
            estilos["Normal"],
        ),
        Spacer(1, 12),
    ]

    # Envolver el texto largo (resumen) en Paragraph para que ajuste.
    data = [columnas]
    for fila in filas:
        data.append([Paragraph(str(c), celda) for c in fila])

    tabla = Table(data, repeatRows=1, colWidths=[55, 80, 70, 55, 70, 110, 60, 230])
    tabla.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ])
    )
    elementos.append(tabla)
    doc.build(elementos)
    return buf.getvalue()
